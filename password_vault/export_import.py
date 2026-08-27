"""
CSV and Excel export / import helpers.
"""

from __future__ import annotations

import csv
import datetime
import logging
import uuid

from .import_profiles import NATIVE, Profile, detect

try:
    import openpyxl
    import openpyxl.styles
    import openpyxl.utils
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

log = logging.getLogger("PasswordVault")

EXPORT_COLS = ["Title", "Username", "Password", "URL", "Category",
               "Notes", "Color", "Created", "Modified", "Pinned"]

_ENTRY_FIELDS = ["title", "username", "password", "url", "category",
                 "notes", "color", "created_at", "modified_at", "pinned"]

# Column order is the single source of truth for both directions.
_FIELD_MAP = dict(zip(EXPORT_COLS, _ENTRY_FIELDS))

_TRUTHY = {"1", "true", "yes", "y", "on"}

# An import larger than this is a mistake or a hostile file, not a vault.
MAX_IMPORT_ROWS = 20000
MAX_IMPORT_BYTES = 32 * 1024 * 1024

# Column auto-width is cosmetic; measuring every cell of a large export is
# not worth the scan.
_WIDTH_SAMPLE_ROWS = 200


# Leading characters that make Excel, LibreOffice and Sheets treat a cell as
# a formula instead of text. A password legitimately starting with one of
# these would otherwise be executed when the export is opened.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _escape_formula(value):
    """Prefix a formula-triggering value so spreadsheets treat it as text."""
    if isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _unescape_formula(value: str) -> str:
    """Reverse :func:`_escape_formula` so a round-trip returns the original."""
    if value[:1] == "'" and value[1:2] in _FORMULA_TRIGGERS:
        return value[1:]
    return value


def _cell_to_text(value) -> str:
    """Render a spreadsheet cell as text.

    Date cells arrive as ``datetime`` objects; ``str()`` on them yields a
    space-separated form that the age parser cannot read, so emit ISO.
    """
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return str(value)


def _entry_to_row(e: dict) -> list[str]:
    return [_escape_formula(e.get(f, "")) for f in _ENTRY_FIELDS]


def _row_to_entry(d: dict, profile: Profile | None = None) -> dict:
    """Convert one source row into an entry dict.

    Keys are matched case-insensitively, so a hand-edited file still
    imports. *profile* selects the source application's column layout;
    ``None`` means this app's own export.
    """
    profile = profile or NATIVE
    now_iso = datetime.datetime.now().isoformat()
    # One lower-cased view of the row, so every lookup below is a plain
    # dict hit rather than a scan per column.
    lowered = {str(k).strip().lower(): v
               for k, v in d.items() if k is not None}

    raw: dict[str, str] = {}
    for header, field_name in profile.columns.items():
        value = lowered.get(header)
        if value in (None, ""):
            continue
        text = _unescape_formula(str(value))
        # Several sources spread one of our fields over more than one
        # column (Chrome's "note"/"notes"); first non-empty wins.
        raw.setdefault(field_name, text)

    notes = raw.get("notes", "")
    extra_lines = []
    for header, label in profile.extras.items():
        value = lowered.get(header)
        if value in (None, ""):
            continue
        extra_lines.append(f"{label}: {_unescape_formula(str(value))}")
    if extra_lines:
        notes = "\n".join(([notes] if notes else []) + extra_lines)

    return {
        "id": str(uuid.uuid4()),
        "title": raw.get("title", ""),
        "username": raw.get("username", ""),
        "password": raw.get("password", ""),
        "url": raw.get("url", ""),
        "category": raw.get("category", "") or "General",
        "notes": notes,
        "color": raw.get("color", "") or "default",
        "pinned": raw.get("pinned", "").strip().lower() in _TRUTHY,
        "created_at": raw.get("created_at", "") or now_iso,
        # Keep the exported timestamp: overwriting it with "now" made every
        # imported entry look freshly rotated and skewed the age stats.
        "modified_at": raw.get("modified_at", "") or now_iso,
    }


def export_csv(entries: list[dict], filepath: str) -> None:
    """Export entries to a CSV file."""
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(EXPORT_COLS)
        for e in entries:
            w.writerow(_entry_to_row(e))


def export_excel(entries: list[dict], filepath: str) -> bool:
    """Export entries to an Excel (.xlsx) file. Returns False if openpyxl missing."""
    if not HAS_OPENPYXL:
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Passwords"
    ws.append(EXPORT_COLS)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for e in entries:
        ws.append(_entry_to_row(e))
    sample = entries[:_WIDTH_SAMPLE_ROWS]
    for col_idx, (header, field) in enumerate(
            zip(EXPORT_COLS, _ENTRY_FIELDS), start=1):
        widest = max([len(header)]
                     + [len(str(e.get(field, "") or "")) for e in sample])
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(widest + 2, 40)
    wb.save(filepath)
    return True


def import_json(filepath: str) -> list[dict]:
    """Import a Bitwarden JSON export.

    Kept apart from the CSV profiles because it is not a column map: the
    format nests folders, multiple URIs, typed items and per-item custom
    fields. :mod:`password_vault.import_json` does the reading; this only
    completes the rows into full entries.
    """
    from .import_json import load

    now_iso = datetime.datetime.now().isoformat()
    entries = []
    for row in load(filepath):
        if len(entries) >= MAX_IMPORT_ROWS:
            log.warning("JSON import truncated at %d rows.",
                        MAX_IMPORT_ROWS)
            break
        if not (row.get("title") or row.get("password")
                or row.get("notes")):
            continue
        entries.append({
            "id": str(uuid.uuid4()),
            "title": row.get("title", ""),
            "username": row.get("username", ""),
            "password": row.get("password", ""),
            "url": row.get("url", ""),
            "category": row.get("category") or "General",
            "notes": row.get("notes", ""),
            "color": "default",
            "pinned": bool(row.get("pinned")),
            "created_at": now_iso,
            "modified_at": now_iso,
        })
    return entries


def read_headers(filepath: str) -> list[str]:
    """Return the header row of *filepath* without parsing the body.

    The import dialog needs the column names to name a profile before it
    commits to reading thousands of rows under the wrong one.
    """
    if filepath.lower().endswith(".xlsx"):
        if not HAS_OPENPYXL:
            return []
        wb = openpyxl.load_workbook(filepath, read_only=True)
        try:
            rows = wb.active.iter_rows(values_only=True)
            try:
                return [str(h).strip() if h else "" for h in next(rows)]
            except StopIteration:
                return []
        finally:
            wb.close()
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        try:
            return [str(h).strip() for h in next(reader)]
        except StopIteration:
            return []


def import_csv(filepath: str, profile: Profile | None = None) -> list[dict]:
    """Import entries from a CSV file (capped at ``MAX_IMPORT_ROWS``).

    *profile* names the source application's column layout; ``None``
    detects it from the header row.
    """
    entries: list[dict] = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        chosen = profile or detect(reader.fieldnames or [])
        for row in reader:
            if len(entries) >= MAX_IMPORT_ROWS:
                log.warning("CSV import truncated at %d rows.",
                            MAX_IMPORT_ROWS)
                break
            e = _row_to_entry(row, chosen)
            if e["title"] or e["password"]:
                entries.append(e)
    return entries


def import_excel(filepath: str,
                 profile: Profile | None = None) -> list[dict]:
    """Import entries from an Excel file, streaming row by row.

    Capped at ``MAX_IMPORT_ROWS`` so a huge or malformed workbook cannot
    exhaust memory.
    """
    if not HAS_OPENPYXL:
        return []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    try:
        rows = wb.active.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []
        headers = [str(h).strip() if h else "" for h in header_row]
        chosen = profile or detect(headers)
        entries: list[dict] = []
        for row in rows:
            if len(entries) >= MAX_IMPORT_ROWS:
                log.warning("Excel import truncated at %d rows.",
                            MAX_IMPORT_ROWS)
                break
            d = dict(zip(headers, [_cell_to_text(v) for v in row]))
            e = _row_to_entry(d, chosen)
            if e["title"] or e["password"]:
                entries.append(e)
        return entries
    finally:
        wb.close()

